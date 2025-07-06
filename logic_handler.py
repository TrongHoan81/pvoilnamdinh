import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle
from datetime import datetime
import re
import io

# ==============================================================================
# HÀM NHẬN DIỆN LOẠI BẢNG KÊ (Đã ổn định)
# ==============================================================================
def detect_report_type(file_content_bytes):
    """
    Hàm nhận diện loại bảng kê dựa trên các dấu hiệu đặc trưng trong file.
    - Bảng kê POS có chữ "Seri" ở ô B4.
    - Bảng kê HDDT có chữ "số công văn (số tham chiếu)" ở dòng 9.
    """
    try:
        # Sử dụng io.BytesIO để đọc nội dung file từ bộ nhớ
        wb = load_workbook(io.BytesIO(file_content_bytes), data_only=True)
        ws = wb.active
        
        # Kiểm tra cho file POS bằng ô B4
        # Dùng str() và lower() để xử lý an toàn, tránh lỗi nếu ô trống hoặc có kiểu dữ liệu khác
        if ws['B4'].value and 'seri' in str(ws['B4'].value).lower().strip():
            return 'POS'
            
        # Kiểm tra cho file HDDT bằng cách duyệt các ô trong dòng 9
        for cell in ws[9]:
            if cell.value and 'số công văn (số tham chiếu)' in str(cell.value).lower():
                return 'HDDT'
    except Exception:
        # Nếu có bất kỳ lỗi nào khi đọc file (ví dụ: file không hợp lệ), trả về UNKNOWN
        return 'UNKNOWN'
        
    # Nếu không tìm thấy dấu hiệu nào, trả về UNKNOWN
    return 'UNKNOWN'

# ==============================================================================
# KHỐI 1: LOGIC XỬ LÝ BẢNG KÊ POS (ĐÃ CẬP NHẬT THEO FILE MỚI NHẤT)
# ==============================================================================

# --- Các hàm trợ giúp cho POS ---
def _pos_to_float(value):
    """Chuyển đổi giá trị sang kiểu float một cách an toàn cho logic POS."""
    try:
        if isinstance(value, str):
            value = value.replace(",", "").strip()
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def _pos_clean_string(s):
    """Làm sạch chuỗi, loại bỏ khoảng trắng thừa cho logic POS."""
    if s is None:
        return ""
    return re.sub(r'\s+', ' ', str(s)).strip()

def _pos_get_static_data(file_path):
    """Đọc dữ liệu cấu hình tĩnh từ file được chỉ định cho logic POS."""
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        chxd_detail_map = {}
        store_specific_x_lookup = {}
        
        for row_idx in range(4, ws.max_row + 1):
            row_values = [cell.value for cell in ws[row_idx]]
            if len(row_values) < 18: continue
            
            chxd_name = _pos_clean_string(row_values[10])
            if chxd_name:
                chxd_detail_map[chxd_name] = {
                    'g5_val': row_values[15],
                    'h5_val': _pos_clean_string(row_values[17]).lower(),
                    'f5_val_full': _pos_clean_string(row_values[16]),
                    'b5_val': chxd_name
                }
                store_specific_x_lookup[chxd_name] = {
                    "xăng e5 ron 92-ii": row_values[11], "xăng ron 95-iii": row_values[12],
                    "dầu do 0,05s-ii": row_values[13], "dầu do 0,001s-v": row_values[14]
                }
        
        def get_lookup(min_r, max_r, min_c=9, max_c=10):
            return {_pos_clean_string(row[0]).lower(): row[1] for row in ws.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c, values_only=True) if row[0] and row[1]}

        tmt_lookup_table = {k: _pos_to_float(v) for k, v in get_lookup(10, 13).items()}

        wb.close()
        return {
            "lookup_table": get_lookup(4, 7),
            "tmt_lookup_table": tmt_lookup_table, "s_lookup_table": get_lookup(29, 31),
            "t_lookup_regular": get_lookup(33, 35), "t_lookup_tmt": get_lookup(48, 50),
            "v_lookup_table": get_lookup(53, 55), "u_value": ws['J36'].value,
            "chxd_detail_map": chxd_detail_map, "store_specific_x_lookup": store_specific_x_lookup
        }
    except FileNotFoundError:
        raise ValueError(f"Lỗi nghiêm trọng: Không tìm thấy file cấu hình '{file_path}'. Vui lòng đảm bảo file này tồn tại trong thư mục của ứng dụng.")
    except Exception as e:
        # Ném ra lỗi để app.py có thể bắt và hiển thị cho người dùng
        raise ValueError(f"Lỗi khi đọc file cấu hình '{file_path}': {e}")

def _pos_create_excel_buffer(processed_rows):
    """Tạo một file Excel trong bộ nhớ từ dữ liệu POS đã xử lý."""
    if not processed_rows:
        return None

    output_wb = Workbook()
    output_ws = output_wb.active
    headers = ["Mã khách", "Tên khách hàng", "Ngày", "Số hóa đơn", "Ký hiệu", "Diễn giải", "Mã hàng", "Tên mặt hàng", "Đvt", "Mã kho", "Mã vị trí", "Mã lô", "Số lượng", "Giá bán", "Tiền hàng", "Mã nt", "Tỷ giá", "Mã thuế", "Tk nợ", "Tk doanh thu", "Tk giá vốn", "Tk thuế có", "Cục thuế", "Vụ việc", "Bộ phận", "Lsx", "Sản phẩm", "Hợp đồng", "Phí", "Khế ước", "Nhân viên bán", "Tên KH(thuế)", "Địa chỉ (thuế)", "Mã số Thuế", "Nhóm Hàng", "Ghi chú", "Tiền thuế"]
    
    for _ in range(4): output_ws.append([''] * len(headers))
    output_ws.append(headers)
    for r_data in processed_rows: output_ws.append(r_data)

    date_style = NamedStyle(name="date_style", number_format='DD/MM/YYYY')
    text_style = NamedStyle(name='text_style', number_format='@')
    
    for row_index in range(6, output_ws.max_row + 1):
        cell_date = output_ws[f'C{row_index}']
        if isinstance(cell_date.value, str) and '-' in cell_date.value:
            try:
                date_obj = datetime.strptime(cell_date.value, '%Y-%m-%d')
                cell_date.value = date_obj 
                cell_date.style = date_style 
            except (ValueError, TypeError): pass
        elif isinstance(cell_date.value, datetime):
            cell_date.style = date_style
        
        output_ws[f'R{row_index}'].style = text_style

    output_ws.column_dimensions['B'].width = 35
    output_ws.column_dimensions['C'].width = 12
    output_ws.column_dimensions['D'].width = 12
    
    output_buffer = io.BytesIO()
    output_wb.save(output_buffer)
    output_buffer.seek(0)
    
    return output_buffer

def _pos_process_single_row(row, details, selected_chxd):
    """Xử lý một dòng hóa đơn đơn lẻ trong bảng kê POS."""
    headers = ["Mã khách", "Tên khách hàng", "Ngày", "Số hóa đơn", "Ký hiệu", "Diễn giải", "Mã hàng", "Tên mặt hàng", "Đvt", "Mã kho", "Mã vị trí", "Mã lô", "Số lượng", "Giá bán", "Tiền hàng", "Mã nt", "Tỷ giá", "Mã thuế", "Tk nợ", "Tk doanh thu", "Tk giá vốn", "Tk thuế có", "Cục thuế", "Vụ việc", "Bộ phận", "Lsx", "Sản phẩm", "Hợp đồng", "Phí", "Khế ước", "Nhân viên bán", "Tên KH(thuế)", "Địa chỉ (thuế)", "Mã số Thuế", "Nhóm Hàng", "Ghi chú", "Tiền thuế"]
    upsse_row = [''] * len(headers)

    try:
        ma_kh = _pos_clean_string(str(row[4]))
        ten_kh = _pos_clean_string(str(row[5]))
        ngay_hd_raw = row[3]
        so_ct = _pos_clean_string(str(row[1]))
        so_hd = _pos_clean_string(str(row[2]))
        dia_chi_goc = _pos_clean_string(str(row[6]))
        mst_goc = _pos_clean_string(str(row[7]))
        product_name = _pos_clean_string(str(row[8]))
        so_luong = _pos_to_float(row[10])
        don_gia_vat = _pos_to_float(row[11])
        tien_hang_source = _pos_to_float(row[13])
        tien_thue_source = _pos_to_float(row[14])
        ma_thue_percent = _pos_to_float(row[15]) if row[15] is not None else 8.0
    except IndexError:
        raise ValueError("Lỗi đọc cột từ file bảng kê POS. Vui lòng đảm bảo file có đủ các cột từ A đến P.")

    upsse_row[0] = ma_kh if ma_kh and len(ma_kh) <= 9 else details['g5_val']
    upsse_row[1] = ten_kh
    
    if isinstance(ngay_hd_raw, datetime): upsse_row[2] = ngay_hd_raw.strftime('%Y-%m-%d')
    elif isinstance(ngay_hd_raw, str):
        try: upsse_row[2] = datetime.strptime(ngay_hd_raw.split(' ')[0], '%d-%m-%Y').strftime('%Y-%m-%d')
        except (ValueError, TypeError): upsse_row[2] = ngay_hd_raw
    else: upsse_row[2] = ngay_hd_raw

    if details['b5_val'] == "Nguyễn Huệ": upsse_row[3] = f"HN{so_hd[-6:]}"
    elif details['b5_val'] == "Mai Linh": upsse_row[3] = f"MM{so_hd[-6:]}"
    else: upsse_row[3] = f"{so_ct[-2:]}{so_hd[-6:]}"

    upsse_row[4] = f"1{so_ct}" if so_ct else ''
    upsse_row[5] = f"Xuất bán lẻ theo hóa đơn số {upsse_row[3]}"
    upsse_row[6] = details['lookup_table'].get(product_name.lower(), '')
    upsse_row[7], upsse_row[8] = product_name, "Lít"
    upsse_row[9] = details['g5_val']
    upsse_row[12] = so_luong
    
    tmt_value = details['tmt_lookup_table'].get(product_name.lower(), 0.0)
    tax_rate_decimal = ma_thue_percent / 100.0

    upsse_row[13] = round(don_gia_vat / (1 + tax_rate_decimal) - tmt_value, 2)
    upsse_row[14] = tien_hang_source - round(tmt_value * so_luong)
    upsse_row[17] = f'{int(ma_thue_percent):02d}'
    upsse_row[18] = details['s_lookup_table'].get(details['h5_val'], '')
    upsse_row[19] = details['t_lookup_regular'].get(details['h5_val'], '')
    upsse_row[20] = details['u_value']
    upsse_row[21] = details['v_lookup_table'].get(details['h5_val'], '')
    upsse_row[23] = details['store_specific_x_lookup'].get(selected_chxd, {}).get(product_name.lower(), '')
    upsse_row[31] = upsse_row[1]
    upsse_row[32] = mst_goc
    upsse_row[33] = dia_chi_goc
    upsse_row[36] = tien_thue_source - round(so_luong * tmt_value * tax_rate_decimal, 0)
    
    return upsse_row

def _pos_create_tmt_row(original_row, tmt_value, details):
    """Tạo dòng Thuế môi trường cho logic POS."""
    tmt_row = list(original_row)
    ma_thue_for_calc = _pos_to_float(original_row[17])
    tax_rate_decimal = ma_thue_for_calc / 100.0

    tmt_row[6], tmt_row[7], tmt_row[8] = "TMT", "Thuế bảo vệ môi trường", "Lít"
    tmt_row[9] = details['g5_val']
    tmt_row[13] = tmt_value
    tmt_row[14] = round(tmt_value * _pos_to_float(original_row[12]), 0)
    
    tmt_row[18] = details['s_lookup_table'].get(details['h5_val'], '')
    tmt_row[19] = details['t_lookup_tmt'].get(details['h5_val'], '')
    tmt_row[20], tmt_row[21] = details['u_value'], details['v_lookup_table'].get(details['h5_val'], '')
    tmt_row[31] = ""
    tmt_row[36] = round(tmt_value * _pos_to_float(original_row[12]) * tax_rate_decimal, 0)
    for idx in [5, 10, 11, 15, 16, 22, 24, 25, 26, 27, 28, 29, 30, 32, 33, 34, 35]:
        if idx < len(tmt_row): tmt_row[idx] = ''
    return tmt_row

def _pos_add_summary_row(original_source_rows, product_name, details, product_tax, selected_chxd, is_new_price_period=False):
    """Tạo dòng tổng hợp cho khách không lấy hóa đơn trong logic POS."""
    headers = ["Mã khách", "Tên khách hàng", "Ngày", "Số hóa đơn", "Ký hiệu", "Diễn giải", "Mã hàng", "Tên mặt hàng", "Đvt", "Mã kho", "Mã vị trí", "Mã lô", "Số lượng", "Giá bán", "Tiền hàng", "Mã nt", "Tỷ giá", "Mã thuế", "Tk nợ", "Tk doanh thu", "Tk giá vốn", "Tk thuế có", "Cục thuế", "Vụ việc", "Bộ phận", "Lsx", "Sản phẩm", "Hợp đồng", "Phí", "Khế ước", "Nhân viên bán", "Tên KH(thuế)", "Địa chỉ (thuế)", "Mã số Thuế", "Nhóm Hàng", "Ghi chú", "Tiền thuế"]
    new_row = [''] * len(headers)

    total_qty = sum(_pos_to_float(r[10]) for r in original_source_rows)
    total_don_gia_vat_x_qty = sum(_pos_to_float(r[11]) * _pos_to_float(r[10]) for r in original_source_rows)
    total_thanh_tien_source = sum(_pos_to_float(r[13]) for r in original_source_rows)
    total_tien_thue_source = sum(_pos_to_float(r[14]) for r in original_source_rows)

    sample_row = original_source_rows[0]
    ngay_hd_raw = sample_row[3]
    so_ct = _pos_clean_string(str(sample_row[1]))

    new_row[0] = details['g5_val']
    new_row[1] = f"Khách hàng mua {product_name} không lấy hóa đơn"
    if isinstance(ngay_hd_raw, datetime): new_row[2] = ngay_hd_raw.strftime('%Y-%m-%d')
    elif isinstance(ngay_hd_raw, str):
        try: new_row[2] = datetime.strptime(ngay_hd_raw.split(' ')[0], '%d-%m-%Y').strftime('%Y-%m-%d')
        except (ValueError, TypeError): new_row[2] = ngay_hd_raw
    else: new_row[2] = ngay_hd_raw
    new_row[4] = f"1{so_ct}" if so_ct else ''
    
    value_C = _pos_clean_string(new_row[2])
    value_E = _pos_clean_string(new_row[4])

    suffix_d_map = {
        "Xăng E5 RON 92-II": "5" if is_new_price_period else "1",
        "Xăng RON 95-III": "6" if is_new_price_period else "2",
        "Dầu DO 0,05S-II": "7" if is_new_price_period else "3",
        "Dầu DO 0,001S-V": "8" if is_new_price_period else "4"
    }
    
    suffix_d = suffix_d_map.get(product_name, "")
    date_part = ""
    if value_C and len(value_C) >= 10:
        try:
            dt_obj = datetime.strptime(value_C, '%Y-%m-%d')
            date_part = f"{dt_obj.day:02d}{dt_obj.month:02d}"
        except ValueError: pass 
        
    if details['b5_val'] == "Nguyễn Huệ": new_row[3] = f"HNBK{date_part}.{suffix_d}"
    elif details['b5_val'] == "Mai Linh": new_row[3] = f"MMBK{date_part}.{suffix_d}"
    else: new_row[3] = f"{value_E[-2:]}BK{date_part}.{suffix_d}"
    
    new_row[5] = f"Xuất bán lẻ theo hóa đơn số {new_row[3]}"
    new_row[6] = details['lookup_table'].get(product_name.lower(), '')
    new_row[7], new_row[8] = product_name, "Lít"
    new_row[9] = details['g5_val']
    new_row[12] = total_qty
    
    tmt_value = details['tmt_lookup_table'].get(product_name.lower(), 0.0)
    tax_rate_decimal = product_tax / 100.0

    avg_don_gia_vat = total_don_gia_vat_x_qty / total_qty if total_qty > 0 else 0
    new_row[13] = round(avg_don_gia_vat / (1 + tax_rate_decimal) - tmt_value, 2)
    new_row[14] = total_thanh_tien_source - round(tmt_value * total_qty)
    new_row[36] = total_tien_thue_source - round(total_qty * tmt_value * tax_rate_decimal, 0)
    new_row[17] = f'{int(product_tax):02d}'
    
    new_row[18] = details['s_lookup_table'].get(details['h5_val'], '')
    new_row[19] = details['t_lookup_regular'].get(details['h5_val'], '')
    new_row[20] = details['u_value']
    new_row[21] = details['v_lookup_table'].get(details['h5_val'], '')
    new_row[23] = details['store_specific_x_lookup'].get(selected_chxd, {}).get(product_name.lower(), '')
    new_row[31] = f"Khách mua {product_name} không lấy hóa đơn"
    
    return new_row

def _pos_generate_upsse_rows(source_data_rows, static_data, selected_chxd, is_new_price_period=False):
    """Hàm chính để xử lý các dòng từ file POS và tạo ra các dòng cho file UpSSE."""
    chxd_details = static_data["chxd_detail_map"].get(selected_chxd)
    if not chxd_details:
        raise ValueError(f"Không tìm thấy thông tin chi tiết cho CHXD: '{selected_chxd}'")
    
    details = {**static_data, **chxd_details}

    final_rows, all_tmt_rows = [], []
    no_invoice_rows = {p: [] for p in ["Xăng E5 RON 92-II", "Xăng RON 95-III", "Dầu DO 0,05S-II", "Dầu DO 0,001S-V"]}
    product_tax_map = {}
    
    for row_idx, row in enumerate(source_data_rows):
        if not row or row[0] is None: continue
        
        try:
            ten_kh = _pos_clean_string(str(row[5]))
            product_name = _pos_clean_string(str(row[8]))
            ma_thue_percent = _pos_to_float(row[15]) if row[15] is not None else 8.0
        except IndexError:
            raise ValueError(f"Dòng {row_idx + 5} trong file bảng kê POS không đủ cột.")

        if product_name and product_name not in product_tax_map:
            product_tax_map[product_name] = ma_thue_percent
        
        if ten_kh == "Người mua không lấy hóa đơn" and product_name in no_invoice_rows:
            no_invoice_rows[product_name].append(row)
        else:
            upsse_row = _pos_process_single_row(row, details, selected_chxd)
            final_rows.append(upsse_row)
            
            tmt_value = details['tmt_lookup_table'].get(product_name.lower(), 0.0)
            so_luong = _pos_to_float(row[10])
            if tmt_value > 0 and so_luong > 0:
                all_tmt_rows.append(_pos_create_tmt_row(upsse_row, tmt_value, details))

    for product, original_rows in no_invoice_rows.items():
        if original_rows:
            product_tax = product_tax_map.get(product, 8.0)
            summary_row = _pos_add_summary_row(original_rows, product, details, product_tax, selected_chxd, is_new_price_period)
            final_rows.append(summary_row)
            
            tmt_unit = details['tmt_lookup_table'].get(product.lower(), 0)
            if tmt_unit > 0 and _pos_to_float(summary_row[12]) > 0:
                tmt_summary = _pos_create_tmt_row(summary_row, tmt_unit, details)
                tmt_summary[1] = summary_row[1]
                all_tmt_rows.append(tmt_summary)

    final_rows.extend(all_tmt_rows)
    return final_rows

def process_pos_report(file_content_bytes, selected_chxd, price_periods, new_price_invoice_number, **kwargs):
    """Hàm điều phối chính cho việc xử lý bảng kê POS."""
    # SỬA LỖI: Thay đổi tên file cấu hình cứng thành "Data_POS.xlsx"
    static_data = _pos_get_static_data("Data_POS.xlsx")
    
    try:
        bkhd_wb = load_workbook(io.BytesIO(file_content_bytes), data_only=True)
        bkhd_ws = bkhd_wb.active
        chxd_details = static_data["chxd_detail_map"].get(selected_chxd)
        if not chxd_details: raise ValueError(f"Không tìm thấy thông tin cho CHXD: '{selected_chxd}'")
        
        b5_bkhd = _pos_clean_string(str(bkhd_ws['B5'].value))
        f5_norm = _pos_clean_string(chxd_details['f5_val_full'])
        if f5_norm.startswith('1'): f5_norm = f5_norm[1:]
        if f5_norm != b5_bkhd:
            raise ValueError(
                "Lỗi dữ liệu: Mã cửa hàng không khớp. Vui lòng kiểm tra lại.\n\n"
                f"   - Mã trong file Bảng kê tải lên (lấy từ ô B5): '{b5_bkhd}'\n"
                f"   - Mã trong file cấu hình Data.xlsx (cột Q):    '{f5_norm}'\n\n"
            )
        all_source_rows = list(bkhd_ws.iter_rows(min_row=5, values_only=True))
        
        if price_periods == '1':
            processed_rows = _pos_generate_upsse_rows(all_source_rows, static_data, selected_chxd, is_new_price_period=False)
            if not processed_rows: raise ValueError("Không có dữ liệu hợp lệ để xử lý trong file POS tải lên.")
            return _pos_create_excel_buffer(processed_rows)
        else:
            if not new_price_invoice_number:
                raise ValueError("Vui lòng nhập 'Số hóa đơn đầu tiên của giá mới' khi chọn 2 giai đoạn giá.")
            
            split_index = -1
            for i, row in enumerate(all_source_rows):
                if len(row) > 2 and row[2] is not None and _pos_clean_string(str(row[2])) == new_price_invoice_number:
                    split_index = i
                    break
            if split_index == -1: raise ValueError(f"Không tìm thấy số hóa đơn '{new_price_invoice_number}' để chia giai đoạn giá.")
            
            old_price_rows = all_source_rows[:split_index]
            new_price_rows = all_source_rows[split_index:]

            buffer_new = _pos_create_excel_buffer(_pos_generate_upsse_rows(new_price_rows, static_data, selected_chxd, is_new_price_period=True))
            buffer_old = _pos_create_excel_buffer(_pos_generate_upsse_rows(old_price_rows, static_data, selected_chxd, is_new_price_period=False))
            
            if not buffer_new and not buffer_old: raise ValueError("Không có dữ liệu hợp lệ để xử lý trong file POS tải lên.")
            return {'new': buffer_new, 'old': buffer_old}
            
    except Exception as e:
        # Ném lại lỗi để tầng trên (app.py) có thể bắt và hiển thị cho người dùng
        raise e

# ==============================================================================
# KHỐI 2: LOGIC HDDT (ĐÃ ỔN ĐỊNH - KHÔNG THAY ĐỔI)
# ==============================================================================
def process_hddt_report(file_content_bytes, selected_chxd, price_periods, new_price_invoice_number, confirmed_date_str=None):
    
    def _clean_string_hddt(s):
        if s is None: return ""
        cleaned_s = str(s).strip()
        if cleaned_s.startswith("'"): cleaned_s = cleaned_s[1:]
        return re.sub(r'\s+', ' ', cleaned_s)

    def _to_float_hddt(value):
        if value is None: return 0.0
        try:
            return float(str(value).replace(',', '').strip())
        except (ValueError, TypeError): return 0.0

    def _format_tax_code_hddt(raw_vat_value):
        if raw_vat_value is None: return ""
        try:
            s_value = str(raw_vat_value).replace('%', '').strip()
            f_value = float(s_value)
            if 0 < f_value < 1: f_value *= 100
            return f"{round(f_value):02d}"
        except (ValueError, TypeError): return ""
    
    def _create_upsse_workbook_hddt():
        headers = ["Mã khách", "Tên khách hàng", "Ngày", "Số hóa đơn", "Ký hiệu", "Diễn giải", "Mã hàng", "Tên mặt hàng", "Đvt", "Mã kho", "Mã vị trí", "Mã lô", "Số lượng", "Giá bán", "Tiền hàng", "Mã nt", "Tỷ giá", "Mã thuế", "Tk nợ", "Tk doanh thu", "Tk giá vốn", "Tk thuế có", "Cục thuế", "Vụ việc", "Bộ phận", "Lsx", "Sản phẩm", "Hợp đồng", "Phí", "Khế ước", "Nhân viên bán", "Tên KH(thuế)", "Địa chỉ (thuế)", "Mã số Thuế", "Nhóm Hàng", "Ghi chú", "Tiền thuế"]
        wb = Workbook()
        ws = wb.active
        for _ in range(4): ws.append([''] * len(headers))
        ws.append(headers)
        return wb

    def _load_static_data_hddt(data_file_path, mahh_file_path, dskh_file_path):
        try:
            static_data = {}
            wb = load_workbook(data_file_path, data_only=True)
            ws = wb.active
            chxd_list, tk_mk_map, khhd_map, chxd_to_khuvuc_map = [], {}, {}, {}
            vu_viec_map = {}
            vu_viec_headers = [_clean_string_hddt(cell.value) for cell in ws[2][4:9]]
            for row_values in ws.iter_rows(min_row=3, max_col=12, values_only=True):
                chxd_name = _clean_string_hddt(row_values[3])
                if chxd_name:
                    ma_kho, khhd, khu_vuc = _clean_string_hddt(row_values[9]), _clean_string_hddt(row_values[10]), _clean_string_hddt(row_values[11])
                    if chxd_name not in tk_mk_map: chxd_list.append(chxd_name)
                    if ma_kho: tk_mk_map[chxd_name] = ma_kho
                    if khhd: khhd_map[chxd_name] = khhd
                    if khu_vuc: chxd_to_khuvuc_map[chxd_name] = khu_vuc
                    vu_viec_map[chxd_name] = {}
                    vu_viec_data_row = row_values[4:9]
                    for i, header in enumerate(vu_viec_headers):
                        if header:
                            key = "Dầu mỡ nhờn" if i == len(vu_viec_headers) - 1 else header
                            vu_viec_map[chxd_name][key] = _clean_string_hddt(vu_viec_data_row[i])
            if not chxd_list: return None, "Không tìm thấy Tên CHXD nào trong cột D của file Data_HDDT.xlsx."
            static_data.update({"DS_CHXD": chxd_list, "tk_mk": tk_mk_map, "khhd_map": khhd_map, "chxd_to_khuvuc_map": chxd_to_khuvuc_map, "vu_viec_map": vu_viec_map})
            def get_lookup_map(min_r, max_r, min_c=1, max_c=2):
                return {_clean_string_hddt(row[0]): row[1] for row in ws.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c, values_only=True) if row[0] and row[1] is not None}
            phi_bvmt_map_raw = get_lookup_map(10, 13)
            static_data["phi_bvmt_map"] = {_clean_string_hddt(k): _to_float_hddt(v) for k, v in phi_bvmt_map_raw.items()}
            static_data.update({
                "tk_no_map": get_lookup_map(29, 31), "tk_doanh_thu_map": get_lookup_map(33, 35),
                "tk_thue_co_map": get_lookup_map(38, 40), "tk_gia_von_value": ws['B36'].value,
                "tk_no_bvmt_map": get_lookup_map(44, 46), "tk_dt_thue_bvmt_map": get_lookup_map(48, 50),
                "tk_gia_von_bvmt_value": ws['B51'].value, "tk_thue_co_bvmt_map": get_lookup_map(53, 55)
            })
            wb_mahh = load_workbook(mahh_file_path, data_only=True)
            static_data["ma_hang_map"] = {_clean_string_hddt(r[0]): _clean_string_hddt(r[2]) for r in wb_mahh.active.iter_rows(min_row=2, max_col=3, values_only=True) if r[0] and r[2]}
            wb_dskh = load_workbook(dskh_file_path, data_only=True)
            static_data["mst_to_makh_map"] = {_clean_string_hddt(r[2]): _clean_string_hddt(r[3]) for r in wb_dskh.active.iter_rows(min_row=2, max_col=4, values_only=True) if r[2]}
            return static_data, None
        except FileNotFoundError as e: return None, f"Lỗi: Không tìm thấy file cấu hình. Chi tiết: {e.filename}"
        except Exception as e: return None, f"Lỗi khi đọc file cấu hình: {e}"

    def _create_hddt_bvmt_row(original_row, phi_bvmt, static_data, khu_vuc):
        bvmt_row = list(original_row)
        so_luong = _to_float_hddt(original_row[12])
        thue_suat = _to_float_hddt(original_row[17]) / 100.0 if original_row[17] else 0.0
        bvmt_row[6], bvmt_row[7] = "TMT", "Thuế bảo vệ môi trường"
        bvmt_row[13], bvmt_row[14] = phi_bvmt, round(phi_bvmt * so_luong)
        bvmt_row[18] = static_data.get('tk_no_bvmt_map', {}).get(khu_vuc)
        bvmt_row[19] = static_data.get('tk_dt_thue_bvmt_map', {}).get(khu_vuc)
        bvmt_row[20] = static_data.get('tk_gia_von_bvmt_value')
        bvmt_row[21] = static_data.get('tk_thue_co_bvmt_map', {}).get(khu_vuc)
        bvmt_row[36] = round(phi_bvmt * so_luong * thue_suat)
        for i in [5, 31, 32, 33]: bvmt_row[i] = ''
        return bvmt_row

    def _generate_upsse_from_hddt_rows(rows_to_process, static_data, selected_chxd, final_date, summary_suffix_map):
        if not rows_to_process: return None
        khu_vuc, ma_kho = static_data['chxd_to_khuvuc_map'].get(selected_chxd), static_data['tk_mk'].get(selected_chxd)
        tk_no, tk_doanh_thu, tk_gia_von, tk_thue_co = static_data['tk_no_map'].get(khu_vuc), static_data['tk_doanh_thu_map'].get(khu_vuc), static_data['tk_gia_von_value'], static_data['tk_thue_co_map'].get(khu_vuc)
        original_invoice_rows, bvmt_rows, summary_data = [], [], {}
        first_invoice_prefix_source = ""
        for bkhd_row in rows_to_process:
            if _to_float_hddt(bkhd_row[8] if len(bkhd_row) > 8 else None) <= 0: continue
            ten_kh, ten_mat_hang = _clean_string_hddt(bkhd_row[3]), _clean_string_hddt(bkhd_row[6])
            is_anonymous, is_petrol = ("không lấy hóa đơn" in ten_kh.lower()), (ten_mat_hang in static_data['phi_bvmt_map'])
            if not is_anonymous or not is_petrol:
                new_upsse_row = [''] * 37
                new_upsse_row[9], new_upsse_row[1], new_upsse_row[31], new_upsse_row[2] = ma_kho, ten_kh, ten_kh, final_date
                so_hd_goc = str(bkhd_row[19] or '').strip()
                new_upsse_row[3] = f"HN{so_hd_goc[-6:]}" if selected_chxd == "Nguyễn Huệ" else f"{(str(bkhd_row[18] or '').strip())[-2:]}{so_hd_goc[-6:]}"
                new_upsse_row[4] = _clean_string_hddt(bkhd_row[17]) + _clean_string_hddt(bkhd_row[18])
                new_upsse_row[5], new_upsse_row[7], new_upsse_row[6] = f"Xuất bán hàng theo hóa đơn số {new_upsse_row[3]}", ten_mat_hang, static_data['ma_hang_map'].get(ten_mat_hang, '')
                new_upsse_row[8], new_upsse_row[12] = _clean_string_hddt(bkhd_row[10]), round(_to_float_hddt(bkhd_row[8]), 3)
                phi_bvmt = static_data['phi_bvmt_map'].get(ten_mat_hang, 0.0) if is_petrol else 0.0
                new_upsse_row[13] = _to_float_hddt(bkhd_row[9]) - phi_bvmt
                ma_thue = _format_tax_code_hddt(bkhd_row[14])
                new_upsse_row[17] = ma_thue
                thue_suat = _to_float_hddt(ma_thue) / 100.0 if ma_thue else 0.0
                tien_thue_goc, so_luong = _to_float_hddt(bkhd_row[15]), _to_float_hddt(bkhd_row[8])
                tien_thue_phi_bvmt = round(phi_bvmt * so_luong * thue_suat)
                new_upsse_row[36] = round(tien_thue_goc - tien_thue_phi_bvmt)
                new_upsse_row[14] = round(_to_float_hddt(bkhd_row[13]) if not is_petrol else _to_float_hddt(bkhd_row[16]) - tien_thue_goc - round(phi_bvmt * so_luong))
                new_upsse_row[18], new_upsse_row[19], new_upsse_row[20], new_upsse_row[21] = tk_no, tk_doanh_thu, tk_gia_von, tk_thue_co
                chxd_vu_viec_map = static_data['vu_viec_map'].get(selected_chxd, {})
                new_upsse_row[23] = chxd_vu_viec_map.get(ten_mat_hang, chxd_vu_viec_map.get("Dầu mỡ nhờn", ''))
                new_upsse_row[32], mst_khach_hang = _clean_string_hddt(bkhd_row[4]), _clean_string_hddt(bkhd_row[5])
                new_upsse_row[33] = mst_khach_hang
                ma_kh_fast = _clean_string_hddt(bkhd_row[2])
                new_upsse_row[0] = ma_kh_fast if ma_kh_fast and len(ma_kh_fast) < 12 else static_data['mst_to_makh_map'].get(mst_khach_hang, ma_kho)
                original_invoice_rows.append(new_upsse_row)
                if is_petrol: bvmt_rows.append(_create_hddt_bvmt_row(new_upsse_row, phi_bvmt, static_data, khu_vuc))
            else:
                if not first_invoice_prefix_source: first_invoice_prefix_source = str(bkhd_row[18] or '').strip()
                if ten_mat_hang not in summary_data:
                    summary_data[ten_mat_hang] = {'sl': 0, 'thue': 0, 'phai_thu': 0, 'first_data': {'mau_so': _clean_string_hddt(bkhd_row[17]),'ky_hieu': _clean_string_hddt(bkhd_row[18]),'don_gia': _to_float_hddt(bkhd_row[9]),'vat_raw': bkhd_row[14]}}
                summary_data[ten_mat_hang]['sl'] += _to_float_hddt(bkhd_row[8])
                summary_data[ten_mat_hang]['thue'] += _to_float_hddt(bkhd_row[15])
                summary_data[ten_mat_hang]['phai_thu'] += _to_float_hddt(bkhd_row[16])
        prefix = first_invoice_prefix_source[-2:] if len(first_invoice_prefix_source) >= 2 else first_invoice_prefix_source
        for product, data in summary_data.items():
            summary_row = [''] * 37
            first_data, total_sl = data['first_data'], data['sl']
            phi_bvmt = static_data['phi_bvmt_map'].get(product, 0.0)
            ma_thue = _format_tax_code_hddt(first_data['vat_raw'])
            thue_suat = _to_float_hddt(ma_thue) / 100.0 if ma_thue else 0.0
            TDT, TTT = data['phai_thu'], data['thue']
            TH_TMT, TT_TMT = round(phi_bvmt * total_sl), round(phi_bvmt * total_sl * thue_suat)
            TT_goc, TH_goc = TTT - TT_TMT, TDT - TH_TMT - (TTT - TT_TMT) - TT_TMT
            summary_row[0], summary_row[1] = ma_kho, f"Khách hàng mua {product} không lấy hóa đơn"
            summary_row[31], summary_row[2] = summary_row[1], final_date
            summary_row[3] = f"{prefix}BK.{final_date.strftime('%d.%m')}.{summary_suffix_map.get(product, '')}"
            summary_row[4] = first_data['mau_so'] + first_data['ky_hieu']
            summary_row[5] = f"Xuất bán hàng theo hóa đơn số {summary_row[3]}"
            summary_row[7], summary_row[6], summary_row[8], summary_row[9] = product, static_data['ma_hang_map'].get(product, ''), "Lít", ma_kho
            summary_row[12], summary_row[13], summary_row[14], summary_row[17] = round(total_sl, 3), first_data['don_gia'] - phi_bvmt, round(TH_goc), ma_thue
            summary_row[18], summary_row[19], summary_row[20], summary_row[21] = tk_no, tk_doanh_thu, tk_gia_von, tk_thue_co
            summary_row[23] = static_data['vu_viec_map'].get(selected_chxd, {}).get(product, '')
            summary_row[36] = round(TT_goc)
            original_invoice_rows.append(summary_row)
            bvmt_rows.append(_create_hddt_bvmt_row(summary_row, phi_bvmt, static_data, khu_vuc))
        upsse_wb = _create_upsse_workbook_hddt()
        for row_data in original_invoice_rows + bvmt_rows: upsse_wb.active.append(row_data)
        output_buffer = io.BytesIO()
        upsse_wb.save(output_buffer)
        output_buffer.seek(0)
        return output_buffer

    # SỬA LỖI: Gọi đúng tên file cấu hình cho HDDT
    static_data, error = _load_static_data_hddt("Data_HDDT.xlsx", "MaHH.xlsx", "DSKH.xlsx")
    if error: raise ValueError(error)
    bkhd_wb = load_workbook(io.BytesIO(file_content_bytes), data_only=True)
    bkhd_ws = bkhd_wb.active
    final_date = None
    if confirmed_date_str:
        final_date = datetime.strptime(confirmed_date_str, '%Y-%m-%d')
    else:
        unique_dates = set()
        for row in bkhd_ws.iter_rows(min_row=11, values_only=True):
            if _to_float_hddt(row[8] if len(row) > 8 else None) > 0:
                date_val = row[20] if len(row) > 20 else None
                if isinstance(date_val, datetime):
                    unique_dates.add(date_val.date())
                elif isinstance(date_val, (int, float)):
                    try:
                        converted_date_obj = pd.to_datetime(date_val, unit='D', origin='1899-12-30').to_pydatetime()
                        unique_dates.add(converted_date_obj.date())
                    except (ValueError, TypeError): pass
                elif isinstance(date_val, str):
                    try: unique_dates.add(datetime.strptime(date_val, '%d/%m/%Y').date())
                    except ValueError: continue
        if not unique_dates: raise ValueError("Không tìm thấy dữ liệu hóa đơn hợp lệ nào trong file Bảng kê HDDT.")
        if len(unique_dates) > 1: raise ValueError("Công cụ chỉ chạy được khi bạn kết xuất hóa đơn trong 1 ngày duy nhất.")
        the_date = unique_dates.pop()
        if the_date.day > 12: final_date = datetime(the_date.year, the_date.month, the_date.day)
        else:
            date1, date2 = datetime(the_date.year, the_date.month, the_date.day), datetime(the_date.year, the_date.day, the_date.month)
            if date1 != date2: return {'choice_needed': True, 'options': [{'text': date1.strftime('%d/%m/%Y'), 'value': date1.strftime('%Y-%m-%d')}, {'text': date2.strftime('%d/%m/%Y'), 'value': date2.strftime('%Y-%m-%d')}]}
            final_date = date1
    all_rows = list(bkhd_ws.iter_rows(min_row=11, values_only=True))
    if price_periods == '1':
        suffix_map = {"Xăng E5 RON 92-II": "1", "Xăng RON 95-III": "2", "Dầu DO 0,05S-II": "3", "Dầu DO 0,001S-V": "4"}
        return _generate_upsse_from_hddt_rows(all_rows, static_data, selected_chxd, final_date, suffix_map)
    else:
        if not new_price_invoice_number: raise ValueError("Vui lòng nhập 'Số hóa đơn đầu tiên của giá mới'.")
        split_index = -1
        for i, row in enumerate(all_rows):
            if str(row[19] or '').strip() == new_price_invoice_number:
                split_index = i
                break
        if split_index == -1: raise ValueError(f"Không tìm thấy hóa đơn số '{new_price_invoice_number}'.")
        suffix_map_old = {"Xăng E5 RON 92-II": "1", "Xăng RON 95-III": "2", "Dầu DO 0,05S-II": "3", "Dầu DO 0,001S-V": "4"}
        suffix_map_new = {"Xăng E5 RON 92-II": "5", "Xăng RON 95-III": "6", "Dầu DO 0,05S-II": "7", "Dầu DO 0,001S-V": "8"}
        result_old = _generate_upsse_from_hddt_rows(all_rows[:split_index], static_data, selected_chxd, final_date, suffix_map_old)
        result_new = _generate_upsse_from_hddt_rows(all_rows[split_index:], static_data, selected_chxd, final_date, suffix_map_new)
        if not result_old and not result_new: raise ValueError("Không có dữ liệu hợp lệ trong cả hai giai đoạn giá.")
        
        output_dict = {}
        if result_old:
            result_old.seek(0)
            output_dict['old'] = result_old
        if result_new:
            result_new.seek(0)
            output_dict['new'] = result_new
        return output_dict

# ==============================================================================
# HÀM ĐIỀU PHỐI CHÍNH
# ==============================================================================
def process_unified_file(file_content, selected_chxd, price_periods, new_price_invoice_number, confirmed_date_str=None):
    """
    Hàm điều phối chính, gọi hàm xử lý tương ứng dựa trên loại bảng kê.
    """
    file_content_bytes = file_content
    report_type = detect_report_type(file_content_bytes)
    
    if report_type == 'POS':
        return process_pos_report(file_content_bytes, selected_chxd, price_periods, new_price_invoice_number)
    elif report_type == 'HDDT':
        return process_hddt_report(file_content_bytes, selected_chxd, price_periods, new_price_invoice_number, confirmed_date_str=confirmed_date_str)
    else:
        raise ValueError("Không thể tự động nhận diện loại Bảng kê. Vui lòng kiểm tra lại file Excel bạn đã tải lên. File phải là bảng kê kết xuất từ POS hoặc HĐĐT.")
