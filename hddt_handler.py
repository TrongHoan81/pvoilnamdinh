import io
import re
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook

def process_hddt_report(file_content_bytes, selected_chxd, price_periods, new_price_invoice_number, confirmed_date_str=None):
    
    # --- CÁC HÀM NỘI BỘ ---
    def _clean_string_hddt(s):
        if s is None: return ""
        cleaned_s = str(s).strip()
        if cleaned_s.startswith("'"): cleaned_s = cleaned_s[1:]
        return re.sub(r'\s+', ' ', cleaned_s)

    def _to_float_hddt(value):
        if value is None: return 0.0
        s = str(value).strip()
        if not s: return 0.0
        if ',' in s and '.' in s and s.rfind('.') < s.rfind(','):
            s = s.replace('.', '').replace(',', '.')
        elif ',' in s and '.' in s and s.rfind(',') < s.rfind('.'):
            s = s.replace(',', '')
        elif ',' in s:
            s = s.replace(',', '.')
        try:
            return float(s)
        except (ValueError, TypeError):
            return 0.0

    # --- HÀM MỚI: TÁI TẠO SỐ THẬP PHÂN CHÍNH XÁC ---
    def _reconstruct_decimal(original_value, approximate_value, expected_decimals):
        """
        Đối chiếu giá trị gốc (có thể sai) với giá trị gần đúng để tái tạo lại số chính xác.
        Ví dụ: original_value=12564, approximate_value=12.564 -> 12.564
        """
        try:
            original_float = float(original_value)
            # Nếu giá trị gần đúng là 0 hoặc chúng đã khớp nhau, không cần làm gì
            if approximate_value == 0 or abs(original_float - approximate_value) < 0.01:
                return approximate_value
            
            ratio = original_float / approximate_value
            # Nếu tỉ lệ là ~1000 (cho số lượng) hoặc ~100 (cho đơn giá)
            if (10**expected_decimals * 0.99) < ratio < (10**expected_decimals * 1.01):
                return original_float / (10**expected_decimals)
            
            # Nếu không rơi vào trường hợp nào, trả về giá trị gần đúng
            return approximate_value
        except (ValueError, TypeError, ZeroDivisionError):
            return approximate_value # Trả về giá trị gần đúng nếu có lỗi

    def _format_tax_code_hddt(raw_vat_value):
        if raw_vat_value is None: return ""
        try:
            s_value = str(raw_vat_value).replace('%', '').strip()
            f_value = float(s_value)
            if 0 < f_value < 1: f_value *= 100
            return f"{round(f_value):02d}"
        except (ValueError, TypeError): return ""
    
    def _create_upsse_workbook_hddt():
        headers = ["Mã khách", "Tên khách hàng", "Ngày", "Số hóa đơn", "Ký hiệu", "Diễn giải", "Mã hàng", "Tên mặt hàng", "Đvt", "Mã kho", "Mã vị trí", "Mã lô", "Số lượng", "Giá bán", "Tiền hàng", "Mã nt", "Tỷ giá", "Mã thuế", "Tk nợ", "Tk doanh thu", "Tk giá vốn", "Tk thuế có", "Cục thuế", "Vụ việc", "Bộ phận", "Lsx", "Sản phẩm", "Hợp đồng", "Phí", "Khế ước", "Nhân viên bán", "Tên KH(thuế)", "Địa chỉ (thuế)", "Mã số Thuế", "Nhóm Hàng", "Ghi chú", "Tiền thuế"]
        wb = Workbook()
        ws = wb.active
        for _ in range(4): ws.append([''] * len(headers))
        ws.append(headers)
        return wb

    def _load_static_data_hddt(data_file_path, mahh_file_path, dskh_file_path):
        try:
            static_data = {}
            wb = load_workbook(data_file_path, data_only=True)
            ws = wb.active
            chxd_list, tk_mk_map, khhd_map, chxd_to_khuvuc_map = [], {}, {}, {}
            vu_viec_map = {}
            vu_viec_headers = [_clean_string_hddt(cell.value) for cell in ws[2][4:9]]
            for row_values in ws.iter_rows(min_row=3, max_col=12, values_only=True):
                chxd_name = _clean_string_hddt(row_values[3])
                if chxd_name:
                    ma_kho, khhd, khu_vuc = _clean_string_hddt(row_values[9]), _clean_string_hddt(row_values[10]), _clean_string_hddt(row_values[11])
                    if chxd_name not in tk_mk_map: chxd_list.append(chxd_name)
                    if ma_kho: tk_mk_map[chxd_name] = ma_kho
                    if khhd: khhd_map[chxd_name] = khhd
                    if khu_vuc: chxd_to_khuvuc_map[chxd_name] = khu_vuc
                    vu_viec_map[chxd_name] = {}
                    vu_viec_data_row = row_values[4:9]
                    for i, header in enumerate(vu_viec_headers):
                        if header:
                            key = "Dầu mỡ nhờn" if i == len(vu_viec_headers) - 1 else header
                            vu_viec_map[chxd_name][key] = _clean_string_hddt(vu_viec_data_row[i])
            if not chxd_list: return None, "Không tìm thấy Tên CHXD nào trong cột D của file Data_HDDT.xlsx."
            static_data.update({"DS_CHXD": chxd_list, "tk_mk": tk_mk_map, "khhd_map": khhd_map, "chxd_to_khuvuc_map": chxd_to_khuvuc_map, "vu_viec_map": vu_viec_map})
            def get_lookup_map(min_r, max_r, min_c=1, max_c=2):
                return {_clean_string_hddt(row[0]): row[1] for row in ws.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c, values_only=True) if row[0] and row[1] is not None}
            phi_bvmt_map_raw = get_lookup_map(10, 13)
            static_data["phi_bvmt_map"] = {_clean_string_hddt(k): _to_float_hddt(v) for k, v in phi_bvmt_map_raw.items()}
            static_data.update({
                "tk_no_map": get_lookup_map(29, 31), "tk_doanh_thu_map": get_lookup_map(33, 35),
                "tk_thue_co_map": get_lookup_map(38, 40), "tk_gia_von_value": ws['B36'].value,
                "tk_no_bvmt_map": get_lookup_map(44, 46), "tk_dt_thue_bvmt_map": get_lookup_map(48, 50),
                "tk_gia_von_bvmt_value": ws['B51'].value, "tk_thue_co_bvmt_map": get_lookup_map(53, 55)
            })
            wb_mahh = load_workbook(mahh_file_path, data_only=True)
            static_data["ma_hang_map"] = {_clean_string_hddt(r[0]): _clean_string_hddt(r[2]) for r in wb_mahh.active.iter_rows(min_row=2, max_col=3, values_only=True) if r[0] and r[2]}
            wb_dskh = load_workbook(dskh_file_path, data_only=True)
            static_data["mst_to_makh_map"] = {_clean_string_hddt(r[2]): _clean_string_hddt(r[3]) for r in wb_dskh.active.iter_rows(min_row=2, max_col=4, values_only=True) if r[2]}
            return static_data, None
        except FileNotFoundError as e: return None, f"Lỗi: Không tìm thấy file cấu hình. Chi tiết: {e.filename}"
        except Exception as e: return None, f"Lỗi khi đọc file cấu hình: {e}"

    def _create_hddt_bvmt_row(original_row, phi_bvmt, static_data, khu_vuc):
        bvmt_row = list(original_row)
        so_luong = original_row[12]
        thue_suat = _to_float_hddt(original_row[17]) / 100.0 if original_row[17] else 0.0
        bvmt_row[6], bvmt_row[7] = "TMT", "Thuế bảo vệ môi trường"
        bvmt_row[13], bvmt_row[14] = phi_bvmt, round(phi_bvmt * so_luong)
        bvmt_row[18] = static_data.get('tk_no_bvmt_map', {}).get(khu_vuc)
        bvmt_row[19] = static_data.get('tk_dt_thue_bvmt_map', {}).get(khu_vuc)
        bvmt_row[20] = static_data.get('tk_gia_von_bvmt_value')
        bvmt_row[21] = static_data.get('tk_thue_co_bvmt_map', {}).get(khu_vuc)
        bvmt_row[36] = round(phi_bvmt * so_luong * thue_suat)
        for i in [5, 31, 32, 33]: bvmt_row[i] = ''
        return bvmt_row

    def _generate_upsse_from_dataframe(df, static_data, selected_chxd, final_date, summary_suffix_map):
        final_date_str = final_date.strftime('%d/%m/%Y')
        khu_vuc, ma_kho = static_data['chxd_to_khuvuc_map'].get(selected_chxd), static_data['tk_mk'].get(selected_chxd)
        tk_no, tk_doanh_thu, tk_gia_von, tk_thue_co = static_data['tk_no_map'].get(khu_vuc), static_data['tk_doanh_thu_map'].get(khu_vuc), static_data['tk_gia_von_value'], static_data['tk_thue_co_map'].get(khu_vuc)
        
        original_invoice_rows, bvmt_rows, summary_data = [], [], {}
        first_invoice_prefix_source = ""

        for index, bkhd_row in df.iterrows():
            # --- *** BẮT ĐẦU KHỐI TÍNH TOÁN NGƯỢC *** ---
            thanh_tien_truoc_thue = _to_float_hddt(bkhd_row.get('thành tiền'))
            tien_thue_gtgt = _to_float_hddt(bkhd_row.get('tiền thuế gtgt'))
            don_gia_raw_val = bkhd_row.get('đơn giá')
            so_luong_raw_val = bkhd_row.get('số lượng')
            ten_mat_hang = _clean_string_hddt(bkhd_row.get('tên hàng hóa, dịch vụ'))
            
            ma_thue = _format_tax_code_hddt(bkhd_row.get('thuế suất'))
            thue_suat = _to_float_hddt(ma_thue) / 100.0 if ma_thue else 0.0
            phi_bvmt = static_data['phi_bvmt_map'].get(ten_mat_hang, 0.0)
            
            don_gia_vat_approx = _to_float_hddt(don_gia_raw_val) / 100.0 if _to_float_hddt(don_gia_raw_val) > 50000 else _to_float_hddt(don_gia_raw_val)
            
            so_luong_approx = 0
            if don_gia_vat_approx > 0 and (1 + thue_suat) > 0:
                don_gia_truoc_thue_co_phi = don_gia_vat_approx / (1 + thue_suat)
                if don_gia_truoc_thue_co_phi > 0:
                    so_luong_approx = thanh_tien_truoc_thue / don_gia_truoc_thue_co_phi
            
            so_luong = _reconstruct_decimal(so_luong_raw_val, so_luong_approx, 3)
            don_gia_vat = _reconstruct_decimal(don_gia_raw_val, don_gia_vat_approx, 2)
            
            if so_luong <= 0: continue
            # --- *** KẾT THÚC KHỐI TÍNH TOÁN NGƯỢC *** ---

            ten_kh = _clean_string_hddt(bkhd_row.get('tên đơn vị'))
            is_anonymous = ("không lấy hóa đơn" in ten_kh.lower())
            is_petrol = (ten_mat_hang in static_data['phi_bvmt_map'])
            
            if not is_anonymous or not is_petrol:
                new_upsse_row = [''] * 37
                new_upsse_row[9], new_upsse_row[1], new_upsse_row[31], new_upsse_row[2] = ma_kho, ten_kh, ten_kh, final_date_str
                so_hd_goc, ky_hieu_hd, mau_so_hd = str(bkhd_row.get('số hóa đơn', '')).strip(), str(bkhd_row.get('ký hiệu hđ', '')).strip(), str(bkhd_row.get('mẫu số hđ', '')).strip()
                new_upsse_row[3] = f"HN{so_hd_goc[-6:]}" if selected_chxd == "Nguyễn Huệ" else f"{ky_hieu_hd[-2:]}{so_hd_goc[-6:]}"
                new_upsse_row[4] = mau_so_hd + ky_hieu_hd
                new_upsse_row[5], new_upsse_row[7], new_upsse_row[6] = f"Xuất bán hàng theo hóa đơn số {new_upsse_row[3]}", ten_mat_hang, static_data['ma_hang_map'].get(ten_mat_hang, '')
                new_upsse_row[8] = _clean_string_hddt(bkhd_row.get('đơn vị tính'))
                
                new_upsse_row[12] = round(so_luong, 3)
                new_upsse_row[13] = (don_gia_vat / (1 + thue_suat)) - phi_bvmt if (1 + thue_suat) > 0 else 0
                new_upsse_row[14] = round(thanh_tien_truoc_thue - (phi_bvmt * so_luong))
                new_upsse_row[17] = ma_thue
                tien_thue_phi_bvmt = round(phi_bvmt * so_luong * thue_suat)
                new_upsse_row[36] = round(tien_thue_gtgt - tien_thue_phi_bvmt)
                
                new_upsse_row[18], new_upsse_row[19], new_upsse_row[20], new_upsse_row[21] = tk_no, tk_doanh_thu, tk_gia_von, tk_thue_co
                chxd_vu_viec_map = static_data['vu_viec_map'].get(selected_chxd, {})
                new_upsse_row[23] = chxd_vu_viec_map.get(ten_mat_hang, chxd_vu_viec_map.get("Dầu mỡ nhờn", ''))
                new_upsse_row[32], mst_khach_hang = _clean_string_hddt(bkhd_row.get('địa chỉ')), _clean_string_hddt(bkhd_row.get('mã số thuế'))
                new_upsse_row[33] = mst_khach_hang
                ma_kh_fast = _clean_string_hddt(bkhd_row.get('mã khách'))
                new_upsse_row[0] = ma_kh_fast if ma_kh_fast and len(ma_kh_fast) < 12 else static_data['mst_to_makh_map'].get(mst_khach_hang, ma_kho)
                original_invoice_rows.append(new_upsse_row)
                if is_petrol: bvmt_rows.append(_create_hddt_bvmt_row(new_upsse_row, phi_bvmt, static_data, khu_vuc))
            else:
                if not first_invoice_prefix_source: first_invoice_prefix_source = str(bkhd_row.get('ký hiệu hđ', '')).strip()
                if ten_mat_hang not in summary_data:
                    summary_data[ten_mat_hang] = {'sl': 0, 'thue': 0, 'phai_thu': 0, 'first_data': {'mau_so': _clean_string_hddt(bkhd_row.get('mẫu số hđ')),'ky_hieu': _clean_string_hddt(bkhd_row.get('ký hiệu hđ')),'don_gia': don_gia_vat,'vat_raw': bkhd_row.get('thuế suất')}}
                summary_data[ten_mat_hang]['sl'] += so_luong
                summary_data[ten_mat_hang]['thue'] += _to_float_hddt(bkhd_row.get('tiền thuế gtgt'))
                summary_data[ten_mat_hang]['phai_thu'] += _to_float_hddt(bkhd_row.get('cộng tiền thanh toán'))

        prefix = first_invoice_prefix_source[-2:] if len(first_invoice_prefix_source) >= 2 else first_invoice_prefix_source
        for product, data in summary_data.items():
            summary_row = [''] * 37
            first_data, total_sl = data['first_data'], data['sl']
            phi_bvmt = static_data['phi_bvmt_map'].get(product, 0.0)
            ma_thue = _format_tax_code_hddt(first_data['vat_raw'])
            thue_suat = _to_float_hddt(ma_thue) / 100.0 if ma_thue else 0.0
            TDT, TTT = data['phai_thu'], data['thue']
            TH_TMT, TT_TMT = round(phi_bvmt * total_sl), round(phi_bvmt * total_sl * thue_suat)
            TT_goc, TH_goc = TTT - TT_TMT, TDT - TH_TMT - (TTT - TT_TMT) - TT_TMT
            summary_row[0], summary_row[1] = ma_kho, f"Khách hàng mua {product} không lấy hóa đơn"
            summary_row[31], summary_row[2] = summary_row[1], final_date_str
            summary_row[3] = f"{prefix}BK.{final_date.strftime('%d.%m')}.{summary_suffix_map.get(product, '')}"
            summary_row[4] = first_data['mau_so'] + first_data['ky_hieu']
            summary_row[5] = f"Xuất bán hàng theo hóa đơn số {summary_row[3]}"
            summary_row[7], summary_row[6], summary_row[8], summary_row[9] = product, static_data['ma_hang_map'].get(product, ''), "Lít", ma_kho
            don_gia_truoc_thue_co_phi = first_data['don_gia'] / (1 + thue_suat) if (1 + thue_suat) > 0 else 0
            summary_row[12], summary_row[13], summary_row[14], summary_row[17] = round(total_sl, 3), don_gia_truoc_thue_co_phi - phi_bvmt, round(TH_goc), ma_thue
            summary_row[18], summary_row[19], summary_row[20], summary_row[21] = tk_no, tk_doanh_thu, tk_gia_von, tk_thue_co
            summary_row[23] = static_data['vu_viec_map'].get(selected_chxd, {}).get(product, '')
            summary_row[36] = round(TT_goc)
            original_invoice_rows.append(summary_row)
            bvmt_rows.append(_create_hddt_bvmt_row(summary_row, phi_bvmt, static_data, khu_vuc))
        
        upsse_wb = _create_upsse_workbook_hddt()
        for row_data in original_invoice_rows + bvmt_rows: upsse_wb.active.append(row_data)
        output_buffer = io.BytesIO()
        upsse_wb.save(output_buffer)
        output_buffer.seek(0)
        return output_buffer

    # --- BẮT ĐẦU LOGIC CHÍNH ---
    static_data, error = _load_static_data_hddt("Data_HDDT.xlsx", "MaHH.xlsx", "DSKH.xlsx")
    if error: raise ValueError(error)

    try:
        df = pd.read_excel(io.BytesIO(file_content_bytes), header=9)
        df.columns = [str(col).strip().lower() for col in df.columns]
    except Exception as e:
        raise ValueError(f"Không thể đọc file Bảng kê. Hãy đảm bảo file đúng định dạng và tiêu đề nằm ở dòng 10. Lỗi: {e}")

    required_cols = ['số lượng', 'tên đơn vị', 'tên hàng hóa, dịch vụ', 'đơn giá', 'thành tiền', 'tiền thuế gtgt', 'cộng tiền thanh toán', 'số hóa đơn']
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"Không tìm thấy cột bắt buộc '{col}' trong file Bảng kê. Vui lòng kiểm tra lại file đầu vào.")

    final_date = None
    if confirmed_date_str:
        final_date = datetime.strptime(confirmed_date_str, '%Y-%m-%d')
    else:
        unique_dates = set()
        date_col_name = next((col for col in df.columns if 'ngày hóa đơn' in col), None)
        if not date_col_name:
            raise ValueError("Không tìm thấy cột 'Ngày hóa đơn' trong file Bảng kê.")
        
        for date_val in df[date_col_name]:
            if pd.notna(date_val):
                if isinstance(date_val, datetime):
                    unique_dates.add(date_val.date())
                elif isinstance(date_val, str):
                    try:
                        unique_dates.add(datetime.strptime(date_val, '%d/%m/%Y').date())
                    except ValueError:
                        continue
        
        if not unique_dates: raise ValueError("Không tìm thấy dữ liệu ngày hợp lệ nào trong file Bảng kê HDDT.")
        if len(unique_dates) > 1: raise ValueError("Công cụ chỉ chạy được khi bạn kết xuất hóa đơn trong 1 ngày duy nhất.")
        the_date = unique_dates.pop()
        
        if the_date.day > 12:
            final_date = datetime(the_date.year, the_date.month, the_date.day)
        else:
            date1 = datetime(the_date.year, the_date.month, the_date.day)
            date2 = datetime(the_date.year, the_date.day, the_date.month)
            if date1 != date2:
                return {'choice_needed': True, 'options': [{'text': date1.strftime('%d/%m/%Y'), 'value': date1.strftime('%Y-%m-%d')}, {'text': date2.strftime('%d/%m/%Y'), 'value': date2.strftime('%Y-%m-%d')}]}
            final_date = date1

    if price_periods == '1':
        suffix_map = {"Xăng E5 RON 92-II": "1", "Xăng RON 95-III": "2", "Dầu DO 0,05S-II": "3", "Dầu DO 0,001S-V": "4"}
        return _generate_upsse_from_dataframe(df, static_data, selected_chxd, final_date, suffix_map)
    else:
        if not new_price_invoice_number: raise ValueError("Vui lòng nhập 'Số hóa đơn đầu tiên của giá mới'.")
        
        split_indices = df.index[df['số hóa đơn'].astype(str).str.strip() == new_price_invoice_number].tolist()
        if not split_indices:
            raise ValueError(f"Không tìm thấy hóa đơn số '{new_price_invoice_number}'.")
        split_index = split_indices[0]

        df_old = df.loc[:split_index-1]
        df_new = df.loc[split_index:]
        
        suffix_map_old = {"Xăng E5 RON 92-II": "1", "Xăng RON 95-III": "2", "Dầu DO 0,05S-II": "3", "Dầu DO 0,001S-V": "4"}
        suffix_map_new = {"Xăng E5 RON 92-II": "5", "Xăng RON 95-III": "6", "Dầu DO 0,05S-II": "7", "Dầu DO 0,001S-V": "8"}
        
        result_old = _generate_upsse_from_dataframe(df_old, static_data, selected_chxd, final_date, suffix_map_old)
        result_new = _generate_upsse_from_dataframe(df_new, static_data, selected_chxd, final_date, suffix_map_new)
        
        if not result_old and not result_new: raise ValueError("Không có dữ liệu hợp lệ trong cả hai giai đoạn giá.")
        
        output_dict = {}
        if result_old:
            output_dict['old'] = result_old
        if result_new:
            output_dict['new'] = result_new
        return output_dict
