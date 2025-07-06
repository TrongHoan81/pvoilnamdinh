import io
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle

# --- Các hàm trợ giúp cho POS ---
def _pos_to_float(value):
    try:
        if isinstance(value, str):
            value = value.replace(",", "").strip()
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def _pos_clean_string(s):
    if s is None:
        return ""
    return re.sub(r'\s+', ' ', str(s)).strip()

def _pos_get_static_data(file_path):
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        chxd_detail_map, store_specific_x_lookup = {}, {}
        for row_idx in range(3, ws.max_row + 1):
            row_values = [cell.value for cell in ws[row_idx]]
            if len(row_values) < 12: continue
            chxd_name = _pos_clean_string(row_values[3])
            if chxd_name:
                chxd_detail_map[chxd_name] = {
                    'g5_val': row_values[9], 'h5_val': _pos_clean_string(row_values[11]).lower(),
                    'f5_val_full': _pos_clean_string(row_values[10]), 'b5_val': chxd_name
                }
                store_specific_x_lookup[chxd_name] = {
                    "xăng e5 ron 92-ii": row_values[4], "xăng ron 95-iii": row_values[5],
                    "dầu do 0,05s-ii": row_values[6], "dầu do 0,001s-v": row_values[7]
                }
        def get_lookup(min_r, max_r, min_c=1, max_c=2):
            return {_pos_clean_string(row[0]).lower(): row[1] for row in ws.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c, values_only=True) if row[0] and row[1]}
        tmt_lookup_table = {k: _pos_to_float(v) for k, v in get_lookup(10, 13).items()}
        wb.close()
        return {
            "lookup_table": get_lookup(4, 7), "tmt_lookup_table": tmt_lookup_table,
            "s_lookup_table": get_lookup(29, 31), "t_lookup_regular": get_lookup(33, 35),
            "t_lookup_tmt": get_lookup(48, 50), "v_lookup_table": get_lookup(53, 55),
            "u_value": ws['B36'].value, "chxd_detail_map": chxd_detail_map,
            "store_specific_x_lookup": store_specific_x_lookup
        }
    except FileNotFoundError:
        raise ValueError(f"Lỗi nghiêm trọng: Không tìm thấy file cấu hình '{file_path}'.")
    except Exception as e:
        raise ValueError(f"Lỗi khi đọc file cấu hình '{file_path}': {e}")

def _pos_create_excel_buffer(processed_rows):
    if not processed_rows: return None
    output_wb = Workbook()
    output_ws = output_wb.active
    headers = ["Mã khách", "Tên khách hàng", "Ngày", "Số hóa đơn", "Ký hiệu", "Diễn giải", "Mã hàng", "Tên mặt hàng", "Đvt", "Mã kho", "Mã vị trí", "Mã lô", "Số lượng", "Giá bán", "Tiền hàng", "Mã nt", "Tỷ giá", "Mã thuế", "Tk nợ", "Tk doanh thu", "Tk giá vốn", "Tk thuế có", "Cục thuế", "Vụ việc", "Bộ phận", "Lsx", "Sản phẩm", "Hợp đồng", "Phí", "Khế ước", "Nhân viên bán", "Tên KH(thuế)", "Địa chỉ (thuế)", "Mã số Thuế", "Nhóm Hàng", "Ghi chú", "Tiền thuế"]
    for _ in range(4): output_ws.append([''] * len(headers))
    output_ws.append(headers)
    for r_data in processed_rows: output_ws.append(r_data)
    text_style = NamedStyle(name='text_style', number_format='@')
    for row_index in range(6, output_ws.max_row + 1):
        output_ws[f'R{row_index}'].style = text_style
    output_ws.column_dimensions['B'].width = 35
    output_ws.column_dimensions['C'].width = 12
    output_ws.column_dimensions['D'].width = 12
    output_buffer = io.BytesIO()
    output_wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer

def _pos_process_single_row(row, details, selected_chxd):
    upsse_row = [''] * 37
    try:
        ma_kh, ten_kh, ngay_hd_raw, so_ct, so_hd, dia_chi_goc, mst_goc, product_name, so_luong, don_gia_vat, tien_hang_source, tien_thue_source = \
        _pos_clean_string(str(row[4])), _pos_clean_string(str(row[5])), row[3], _pos_clean_string(str(row[1])), _pos_clean_string(str(row[2])), \
        _pos_clean_string(str(row[6])), _pos_clean_string(str(row[7])), _pos_clean_string(str(row[8])), _pos_to_float(row[10]), \
        _pos_to_float(row[11]), _pos_to_float(row[13]), _pos_to_float(row[14])
        ma_thue_percent = _pos_to_float(row[15]) if row[15] is not None else 8.0
    except IndexError:
        raise ValueError("Lỗi đọc cột từ file bảng kê POS. Vui lòng đảm bảo file có đủ các cột từ A đến P.")
    
    upsse_row[0] = ma_kh if ma_kh and len(ma_kh) <= 9 else details['g5_val']
    upsse_row[1] = ten_kh
    
    if isinstance(ngay_hd_raw, datetime):
        upsse_row[2] = ngay_hd_raw.strftime('%d/%m/%Y')
    else:
        upsse_row[2] = str(ngay_hd_raw).split(' ')[0]

    if details['b5_val'] == "Nguyễn Huệ": upsse_row[3] = f"HN{so_hd[-6:]}"
    elif details['b5_val'] == "Mai Linh": upsse_row[3] = f"MM{so_hd[-6:]}"
    else: upsse_row[3] = f"{so_ct[-2:]}{so_hd[-6:]}"
    
    upsse_row[4] = f"1{so_ct}" if so_ct else ''
    upsse_row[5] = f"Xuất bán lẻ theo hóa đơn số {upsse_row[3]}"
    upsse_row[6] = details['lookup_table'].get(product_name.lower(), '')
    upsse_row[7], upsse_row[8] = product_name, "Lít"
    upsse_row[9] = details['g5_val']
    upsse_row[12] = so_luong
    tmt_value = details['tmt_lookup_table'].get(product_name.lower(), 0.0)
    tax_rate_decimal = ma_thue_percent / 100.0
    upsse_row[13] = round(don_gia_vat / (1 + tax_rate_decimal) - tmt_value, 2)
    upsse_row[14] = tien_hang_source - round(tmt_value * so_luong)
    upsse_row[17] = f'{int(ma_thue_percent):02d}'
    upsse_row[18] = details['s_lookup_table'].get(details['h5_val'], '')
    upsse_row[19] = details['t_lookup_regular'].get(details['h5_val'], '')
    upsse_row[20] = details['u_value']
    upsse_row[21] = details['v_lookup_table'].get(details['h5_val'], '')
    upsse_row[23] = details['store_specific_x_lookup'].get(selected_chxd, {}).get(product_name.lower(), '')
    upsse_row[31] = upsse_row[1]
    upsse_row[32] = mst_goc
    upsse_row[33] = dia_chi_goc
    upsse_row[36] = tien_thue_source - round(so_luong * tmt_value * tax_rate_decimal, 0)
    return upsse_row

def _pos_create_tmt_row(original_row, tmt_value, details):
    tmt_row = list(original_row)
    ma_thue_for_calc = _pos_to_float(original_row[17])
    tax_rate_decimal = ma_thue_for_calc / 100.0
    tmt_row[6], tmt_row[7], tmt_row[8] = "TMT", "Thuế bảo vệ môi trường", "Lít"
    tmt_row[9] = details['g5_val']
    tmt_row[13] = tmt_value
    tmt_row[14] = round(tmt_value * _pos_to_float(original_row[12]), 0)
    tmt_row[18] = details['s_lookup_table'].get(details['h5_val'], '')
    tmt_row[19] = details['t_lookup_tmt'].get(details['h5_val'], '')
    tmt_row[20], tmt_row[21] = details['u_value'], details['v_lookup_table'].get(details['h5_val'], '')
    tmt_row[31] = ""
    tmt_row[36] = round(tmt_value * _pos_to_float(original_row[12]) * tax_rate_decimal, 0)
    for idx in [5, 10, 11, 15, 16, 22, 24, 25, 26, 27, 28, 29, 30, 32, 33, 34, 35]:
        if idx < len(tmt_row): tmt_row[idx] = ''
    return tmt_row

def _pos_add_summary_row(original_source_rows, product_name, details, product_tax, selected_chxd, is_new_price_period=False):
    new_row = [''] * 37
    total_qty = sum(_pos_to_float(r[10]) for r in original_source_rows)
    total_don_gia_vat_x_qty = sum(_pos_to_float(r[11]) * _pos_to_float(r[10]) for r in original_source_rows)
    total_thanh_tien_source = sum(_pos_to_float(r[13]) for r in original_source_rows)
    total_tien_thue_source = sum(_pos_to_float(r[14]) for r in original_source_rows)
    sample_row = original_source_rows[0]
    ngay_hd_raw = sample_row[3]
    so_ct = _pos_clean_string(str(sample_row[1]))
    new_row[0] = details['g5_val']
    new_row[1] = f"Khách hàng mua {product_name} không lấy hóa đơn"

    if isinstance(ngay_hd_raw, datetime):
        new_row[2] = ngay_hd_raw.strftime('%d/%m/%Y')
    else:
        new_row[2] = str(ngay_hd_raw).split(' ')[0]

    new_row[4] = f"1{so_ct}" if so_ct else ''
    value_E = _pos_clean_string(new_row[4])
    suffix_d_map = {"Xăng E5 RON 92-II": "5" if is_new_price_period else "1", "Xăng RON 95-III": "6" if is_new_price_period else "2", "Dầu DO 0,05S-II": "7" if is_new_price_period else "3", "Dầu DO 0,001S-V": "8" if is_new_price_period else "4"}
    suffix_d = suffix_d_map.get(product_name, "")
    date_part = ""
    try:
        dt_obj = datetime.strptime(new_row[2], '%d/%m/%Y')
        date_part = f"{dt_obj.day:02d}{dt_obj.month:02d}"
    except (ValueError, TypeError):
        pass
    
    if details['b5_val'] == "Nguyễn Huệ": new_row[3] = f"HNBK{date_part}.{suffix_d}"
    elif details['b5_val'] == "Mai Linh": new_row[3] = f"MMBK{date_part}.{suffix_d}"
    else: new_row[3] = f"{value_E[-2:]}BK{date_part}.{suffix_d}"

    new_row[5] = f"Xuất bán lẻ theo hóa đơn số {new_row[3]}"
    new_row[6] = details['lookup_table'].get(product_name.lower(), '')
    new_row[7], new_row[8] = product_name, "Lít"
    new_row[9] = details['g5_val']
    new_row[12] = total_qty
    tmt_value = details['tmt_lookup_table'].get(product_name.lower(), 0.0)
    tax_rate_decimal = product_tax / 100.0
    avg_don_gia_vat = total_don_gia_vat_x_qty / total_qty if total_qty > 0 else 0
    new_row[13] = round(avg_don_gia_vat / (1 + tax_rate_decimal) - tmt_value, 2)
    new_row[14] = total_thanh_tien_source - round(tmt_value * total_qty)
    new_row[36] = total_tien_thue_source - round(total_qty * tmt_value * tax_rate_decimal, 0)
    new_row[17] = f'{int(product_tax):02d}'
    new_row[18] = details['s_lookup_table'].get(details['h5_val'], '')
    new_row[19] = details['t_lookup_regular'].get(details['h5_val'], '')
    new_row[20] = details['u_value']
    new_row[21] = details['v_lookup_table'].get(details['h5_val'], '')
    new_row[23] = details['store_specific_x_lookup'].get(selected_chxd, {}).get(product_name.lower(), '')
    new_row[31] = f"Khách mua {product_name} không lấy hóa đơn"
    return new_row

def _pos_generate_upsse_rows(source_data_rows, static_data, selected_chxd, is_new_price_period=False):
    chxd_details = static_data["chxd_detail_map"].get(selected_chxd)
    if not chxd_details: raise ValueError(f"Không tìm thấy thông tin chi tiết cho CHXD: '{selected_chxd}'")
    details = {**static_data, **chxd_details}
    final_rows, all_tmt_rows = [], []
    no_invoice_rows = {p: [] for p in ["Xăng E5 RON 92-II", "Xăng RON 95-III", "Dầu DO 0,05S-II", "Dầu DO 0,001S-V"]}
    product_tax_map = {}
    for row_idx, row in enumerate(source_data_rows):
        if not row or row[0] is None: continue
        try:
            ten_kh, product_name, ma_thue_percent = _pos_clean_string(str(row[5])), _pos_clean_string(str(row[8])), _pos_to_float(row[15]) if row[15] is not None else 8.0
        except IndexError: raise ValueError(f"Dòng {row_idx + 5} trong file bảng kê POS không đủ cột.")
        if product_name and product_name not in product_tax_map: product_tax_map[product_name] = ma_thue_percent
        if ten_kh == "Người mua không lấy hóa đơn" and product_name in no_invoice_rows:
            no_invoice_rows[product_name].append(row)
        else:
            upsse_row = _pos_process_single_row(row, details, selected_chxd)
            final_rows.append(upsse_row)
            tmt_value = details['tmt_lookup_table'].get(product_name.lower(), 0.0)
            so_luong = _pos_to_float(row[10])
            if tmt_value > 0 and so_luong > 0:
                all_tmt_rows.append(_pos_create_tmt_row(upsse_row, tmt_value, details))
    for product, original_rows in no_invoice_rows.items():
        if original_rows:
            product_tax = product_tax_map.get(product, 8.0)
            summary_row = _pos_add_summary_row(original_rows, product, details, product_tax, selected_chxd, is_new_price_period)
            final_rows.append(summary_row)
            tmt_unit = details['tmt_lookup_table'].get(product.lower(), 0)
            if tmt_unit > 0 and _pos_to_float(summary_row[12]) > 0:
                tmt_summary = _pos_create_tmt_row(summary_row, tmt_unit, details)
                tmt_summary[1] = summary_row[1]
                all_tmt_rows.append(tmt_summary)
    final_rows.extend(all_tmt_rows)
    return final_rows

def process_pos_report(file_content_bytes, selected_chxd, price_periods, new_price_invoice_number, **kwargs):
    static_data = _pos_get_static_data("Data_HDDT.xlsx")
    try:
        bkhd_wb = load_workbook(io.BytesIO(file_content_bytes), data_only=True)
        bkhd_ws = bkhd_wb.active
        chxd_details = static_data["chxd_detail_map"].get(selected_chxd)
        if not chxd_details: raise ValueError(f"Không tìm thấy thông tin cho CHXD: '{selected_chxd}'")
        b5_bkhd = _pos_clean_string(str(bkhd_ws['B5'].value))
        f5_norm = _pos_clean_string(chxd_details['f5_val_full'])
        if f5_norm and len(f5_norm) >= 6 and f5_norm[-6:] != b5_bkhd:
            raise ValueError(f"Lỗi dữ liệu: Mã cửa hàng không khớp.\n- Mã trong Bảng kê (ô B5): '{b5_bkhd}'\n- Mã trong file cấu hình (6 ký tự cuối cột K): '{f5_norm[-6:]}'")
        all_source_rows = list(bkhd_ws.iter_rows(min_row=5, values_only=True))
        if price_periods == '1':
            processed_rows = _pos_generate_upsse_rows(all_source_rows, static_data, selected_chxd, is_new_price_period=False)
            if not processed_rows: raise ValueError("Không có dữ liệu hợp lệ để xử lý trong file POS tải lên.")
            return _pos_create_excel_buffer(processed_rows)
        else:
            if not new_price_invoice_number: raise ValueError("Vui lòng nhập 'Số hóa đơn đầu tiên của giá mới' khi chọn 2 giai đoạn giá.")
            split_index = -1
            for i, row in enumerate(all_source_rows):
                if len(row) > 2 and row[2] is not None and _pos_clean_string(str(row[2])) == new_price_invoice_number:
                    split_index = i
                    break
            if split_index == -1: raise ValueError(f"Không tìm thấy số hóa đơn '{new_price_invoice_number}' để chia giai đoạn giá.")
            old_price_rows, new_price_rows = all_source_rows[:split_index], all_source_rows[split_index:]
            buffer_new = _pos_create_excel_buffer(_pos_generate_upsse_rows(new_price_rows, static_data, selected_chxd, is_new_price_period=True))
            buffer_old = _pos_create_excel_buffer(_pos_generate_upsse_rows(old_price_rows, static_data, selected_chxd, is_new_price_period=False))
            if not buffer_new and not buffer_old: raise ValueError("Không có dữ liệu hợp lệ để xử lý trong file POS tải lên.")
            return {'new': buffer_new, 'old': buffer_old}
    except Exception as e:
        raise e
