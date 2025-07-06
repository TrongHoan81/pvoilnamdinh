import io
from openpyxl import load_workbook

def detect_report_type(file_content_bytes):
    """
    Hàm nhận diện loại bảng kê dựa trên các dấu hiệu đặc trưng trong file.
    - Bảng kê POS có chữ "Seri" ở ô B4.
    - Bảng kê HDDT có chữ "số công văn (số tham chiếu)" ở dòng 9.
    """
    try:
        wb = load_workbook(io.BytesIO(file_content_bytes), data_only=True)
        ws = wb.active
        
        # Kiểm tra cho file POS bằng ô B4
        if ws['B4'].value and 'seri' in str(ws['B4'].value).lower().strip():
            return 'POS'
            
        # Kiểm tra cho file HDDT bằng cách duyệt các ô trong dòng 9
        for cell in ws[9]:
            if cell.value and 'số công văn (số tham chiếu)' in str(cell.value).lower():
                return 'HDDT'
    except Exception:
        # Nếu có bất kỳ lỗi nào khi đọc file (ví dụ: file không hợp lệ), trả về UNKNOWN
        return 'UNKNOWN'
        
    # Nếu không tìm thấy dấu hiệu nào, trả về UNKNOWN
    return 'UNKNOWN'
